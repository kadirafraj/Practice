import openpyxl


class LoginPageData:
    @staticmethod
    def gettestdata(test_case_name):
        book=openpyxl.load_workbook("C:\\Users\\afraj\\Desktop\\Facebook.xlsx")
        sheet=book.active
        dic={}
        for i in range(2,sheet.max_row+1):
            if sheet.cell(row=i, column=1).value==test_case_name:
                for j in range(2, sheet.max_column+1):
                    dic[sheet.cell(row=1, column=j).value]=sheet.cell(row=i, column=j).value
        return[dic]

